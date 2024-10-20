from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import time
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, numbers
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

class MoexDownloader:
    def __init__(self):
        self.url = "https://www.moex.com/"
        self.download_dir = os.path.dirname(os.path.abspath(__file__))
        self.driver = self.setup_driver()
        
        
    def setup_driver(self):
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--headless")
        chrome_options.add_experimental_option("prefs", {
            "download.default_directory": self.download_dir,
            "download.prompt_for_download" : False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        })
        chrome_options.add_experimental_option("excludeSwitches", 
                                               ["enable-logging"])
        
        
        return webdriver.Chrome(options=chrome_options)
    
    
    def wait_for_download(self, file_prefix, timeout=10):
        end_time = time.time() + timeout
        while time.time() < end_time:
            files = os.listdir(self.download_dir)
            matching_files = [f for f in files if f.startswith(file_prefix)]
            if matching_files:
                if not any(fname.endswith('.xml') for fname in matching_files):
                    return os.path.join(self.download_dir, matching_files[0])
                time.sleep(1)
    
    
    def nav_to_indicated_charts(self):
        self.driver.get(self.url)
        
        WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable(
            (By.CLASS_NAME, "header__button"))).click()
        WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable(
            (By.XPATH, "//a[text()='Срочный рынок']"))).click()
        
        try:
            WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable(
                (By.XPATH, "//a[text()='Согласен']"))).click()
            WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable(
                (By.XPATH, "//span[text()='Индикативные курсы']"))).click()
        except:
            pass
        
        
    def download_data(self, prefix, currency):
        
        WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable(
            (By.CLASS_NAME, "ui-select__placeholder"))).click()
        currency_element = WebDriverWait(self.driver, 30).until(
            EC.visibility_of_element_located(
                (By.XPATH, f"//a[text()='{currency}']")))
        
        listbox_container = WebDriverWait(self.driver, 30).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, '[aria-role="listbox"]')))
        
        actions = ActionChains(self.driver)
        actions.move_to_element(listbox_container)
        actions.scroll_to_element(currency_element)
        actions.perform()
        
        WebDriverWait(self.driver, 30).until(EC.visibility_of(currency_element))
        currency_element.click()

        get_data_link = WebDriverWait(self.driver, 30).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//a[text()='Получить данные в XML']")))
        
        get_data_link.click()
        
        return self.wait_for_download(prefix)
    
        
    def parse_xml(self, file_path):
        tree = ET.parse(file_path)
        root = tree.getroot()
        data = {}
        for row in root.findall(".//row"):
            tradedate = row.get('tradedate')
            tradetime = row.get('tradetime')
            secid = row.get('secid')
            rate = float(row.get('rate'))
            clearing = row.get('clearing')
            if tradedate not in data:
                data[tradedate] = {'vk': None, 'pk': None}
            if clearing == 'vk':
                data[tradedate]['vk'] = rate
            elif clearing == 'pk':
                data[tradedate]['pk'] = rate
        return [(date, values['pk'], values['vk']) for date, values in data.items()]
    
    
    def create_excel(self, usd_data, eur_data):
        wb = Workbook()
        ws = wb.active

        ws['A1'] = 'Дата'
        ws['B1'] = 'Значение курса промежуточного клиринга'
        ws['C1'] = 'Значение курса основного клиринга'
        ws['E1'] = 'Дата'
        ws['F1'] = 'Значение курса промежуточного клиринга'
        ws['G1'] = 'Значение курса основного клиринга'
        ws['H1'] = 'Изменение'
        
        row_index = 2
        for data in usd_data:
            ws.cell(row=row_index, column=1, value=data[0])
            ws.cell(row=row_index, column=2, value=data[1])
            ws.cell(row=row_index, column=3, value=data[2])
            row_index += 1
            
        row_index = 2
        
        for data in eur_data:
            ws.cell(row=row_index, column=5, value=data[0])
            ws.cell(row=row_index, column=6, value=data[1])
            ws.cell(row=row_index, column=7, value=data[2])
            row_index += 1
        

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        financial_style = NamedStyle(name="financial_style")
        financial_style.font = Font(name='Calibri', size=11)
        financial_style.alignment = Alignment(horizontal='right')
        financial_style.number_format = '#,##0.00 ₽_);[Red](#,##0.00 ₽)'

        for row in ws.iter_rows(min_row=2, max_col=8, max_row=ws.max_row):
            for cell in row[1:3] + row[5:8]:
                if isinstance(cell.value, (int, float)):
                    cell.style = financial_style
                    
            euro_rate_vk = row[6].value
            usd_rate_vk = row[2].value
            if euro_rate_vk is not None and usd_rate_vk is not None:
                change = euro_rate_vk / usd_rate_vk
                row[7].value = change
                row[7].style = financial_style
            else:
                row[7].value = '-'
        timestamp = time.strftime("%Y%m%d%H%M%S")
        save_path = os.path.join(self.download_dir, f'currency_rates_{timestamp}.xlsx')
        wb.save(save_path)
        
        return save_path, ws.max_row - 1
 
        
    def send_email(self, file_path, row_count):
        from_email = "polzovateltehniceskij@gmail.com"
        to_email = "asmirnov0940@yandex.ru"
        subject = "Отчет о курсах валют"
        body = f"Файл с отчетом о курсах валют прикреплен к письму.\n\nКоличество строк в файле: {row_count} {self.decline_rows(row_count)}."

        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject

        msg.attach(MIMEText(body, 'plain'))

        with open(file_path, "rb") as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename= {os.path.basename(file_path)}')
            msg.attach(part)

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.connect('smtp.gmail.com', 587)
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(from_email, "")
        #server.auth_plain()
        text = msg.as_string()
        server.sendmail(from_email, to_email, text)
        server.quit()


    def decline_rows(self, count):
        if 11 <= count % 100 <= 19:
            return "строк"
        elif count % 10 == 1:
            return "строка"
        elif 2 <= count % 10 <= 4:
            return "строки"
        else:
            return "строк"    
        

    def run(self):
        self.nav_to_indicated_charts()
        
        self.download_data(currency="USD/RUB - Доллар США к российскому рублю", 
                           prefix="currencyRate-USD_RUB")
        usd_file = [f for f in os.listdir(self.download_dir) 
                    if f.startswith('currencyRate-USD_RUB')][0]
        usd_data = self.parse_xml(os.path.join(self.download_dir, usd_file))
        os.remove(os.path.join(self.download_dir, usd_file))
        
        
        self.download_data(currency="EUR/RUB - Евро к российскому рублю", 
                           prefix="currencyRate-EUR_RUB")
        eur_file = [f for f in os.listdir(self.download_dir) 
                    if f.startswith('currencyRate-EUR_RUB')][0]
        eur_data = self.parse_xml(os.path.join(self.download_dir, eur_file))
        os.remove(os.path.join(self.download_dir, eur_file))

        file_path, row_count = self.create_excel(usd_data, eur_data)
        
        self.send_email(file_path, row_count)

        self.driver.quit()


if __name__ == "__main__":
    app = MoexDownloader()
    app.run()
